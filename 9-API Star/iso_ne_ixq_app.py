import chardet
import json
import pandas as pd
import sys
from typing import List
from openpyxl import load_workbook

from apistar import App, Route, types, validators
from apistar.http import JSONResponse
from datetime import date, datetime

def _load_nyiso_data():
    # with open('NYISO-Interconnection-Queue.xlsx') as file:
    file = 'NYISO-Interconnection-Queue.xlsx'
    df_active = pd.read_excel(file, skiprows=1, sheet_name='Active')
    # df_withdrawn = pd.read_excel(file, encoding=sys.getfilesystemencoding(), sheet_name='Withdrawn')

    df_active['id'] = df_active.index.values
    cols = [ 'id', 'Pos.', 'Owner/Developer', 'Project Name', 'of IR', '(MW)', '(MW).1',
       'Fuel', 'County/State', 'Unnamed: 8', 'Point', 'Utility ', 'S',
       'Unnamed: 12', 'of Studies', 'SGIA Tender', ' In-Service',
       'Initial-Sync', 'COD',]
    df_active = df_active[cols]

    df_active.drop(columns=['of IR','Unnamed: 12','SGIA Tender',' In-Service','Initial-Sync','COD'], inplace=True)
    df_active = df_active.fillna('')
    # file = "test.xlsx"  # load the work book
    # wb_obj = load_workbook(filename=file)
    # wsheet = wb_obj['test']
    # dataDict = {}
    #
    # for key, *values in wsheet.iter_rows():
    #     dataDict[key.value] = [v.value for v in values]
    #
    # return {project['id']: project for project in
    return df_active.to_dict('index')
    # return df_active.to_dict('index')

projects = _load_nyiso_data()
VALID_OWNERS = set([project['Owner/Developer'] for project in projects.values()])
PROJECT_NOT_FOUND = 'Project not found in active queue. Try the withdrawn project list?'

class Project(types.Type):
    id = validators.Integer(allow_null=True) # assign in POST
    # owner = validators.String(enum=list(VALID_OWNERS))
    pass
    
    
# API Methods
# list = sorted(projects.items())
def list_projects():
    return [project[1] for project in sorted(projects.items())]

def create_project(project: Project) -> JSONResponse:
    project_id = max(projects.keys())+1
    project.id = project_id
    projects[project_id] = project
    return JSONResponse(Project(project), status_code=201)


def get_project(project_id: int) -> JSONResponse:
    project = projects.get(project_id)
    if not project:
        error = {'error': PROJECT_NOT_FOUND}
        return JSONResponse(error, status_code=404)

    return JSONResponse(project, status_code=200)


def update_project(project_id: int, project: Project) -> JSONResponse:
    if not projects.get(project_id):
        error = {'error': PROJECT_NOT_FOUND}
        return JSONResponse(error, status_code=404)

    project.id = project_id
    projects[project_id] = project
    return JSONResponse(Project(project), status_code=200)


def delete_project(project_id: int) -> JSONResponse:
    if not projects.get(project_id):
        error = {'error': PROJECT_NOT_FOUND}
        return JSONResponse(error, status_code=404)

    del projects[project_id]
    return JSONResponse({}, status_code=204)


routes = [
    Route('/', method='GET', handler=list_projects),
    Route('/', method='POST', handler=create_project),
    Route('/{project_id}/', method='GET', handler=get_project),
    Route('/{project_id}/', method='PUT', handler=update_project),
    Route('/{project_id}/', method='DELETE', handler=delete_project),
]

app = App(routes=routes)


if __name__ == '__main__':
    app.serve('127.0.0.1', 5000, debug=True)

